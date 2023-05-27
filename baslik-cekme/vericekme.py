from xml.etree.ElementTree import tostring
from openpyxl import load_workbook
import pandas as pd
import xlsxwriter
import requests
from bs4 import BeautifulSoup
#satır sayısı
wb = load_workbook("sade-link.xlsx")
sheet = wb.worksheets[0]
row_count = sheet.max_row
#yazdırma işlemleri
workbook = xlsxwriter.Workbook('tum.xlsx')
worksheet = workbook.add_worksheet()
#satır sayısı kadar veri çekip onları excel dosyasına kaydediyor 
for i in range(215):
    #exceldeki linkleri okuyor
    excel_data = pd.read_excel('sade-link.xlsx',index_col=0)
    data = excel_data.index[i-1]
    #belirtilen urldeki bir kısmı çekiyor
    url = data
    response = requests.get(url)
    html_icerigi = response.content
    soup = BeautifulSoup(html_icerigi,"html.parser")
    isim =soup.find_all("h1",{"itemprop":"name"})
    #h1 etiketini silip sadece içeriğini alıyor
    isim[0] = (isim[0].text).strip("\n").strip()
    #siteden çekilen verileri excel tablosuna yazdırıyor
    worksheet.write('B'+ str(i) , isim[0])
    i = int(i)
workbook.close()